# -*- coding: utf-8 -*-
import base64
import csv
import io
from odoo import api, fields, models, _
from odoo.exceptions import UserError
from odoo.addons.queue_job.delay import group, chain

try:
    import openpyxl  # optional, only for .xlsx
except Exception:
    openpyxl = None


class HrEmployeeImportWizard(models.TransientModel):
    _name = "hr.employee.import.wizard"
    _description = "Import Employees Asynchronously"

    file = fields.Binary(required=True)
    file_name = fields.Char(required=True)
    has_header = fields.Boolean(default=True)
    chunk_size = fields.Integer(default=500)
    notify_done = fields.Boolean(default=True, string="Notify by Email when done")

    def _decode_file(self):
        if not self.file or not self.file_name:
            raise UserError(_("Please upload a file."))
        return base64.b64decode(self.file)

    def _iter_rows_csv(self, content):
        """Yield dict rows using header names.

        This is more robust than positional mapping and allows
        arbitrary column order as long as headers are present.
        """
        f = io.StringIO(content.decode("utf-8", errors="replace"))
        # If file has a header, DictReader maps columns by header name.
        # If no header, fall back to simple reader and numeric keys.
        if self.has_header:
            reader = csv.DictReader(f)
            for row in reader:
                yield {k or "": (v or "") for k, v in (row or {}).items()}
        else:
            reader = csv.reader(f)
            for row in reader:
                # produce dict with numeric keys as strings
                yield {str(i): (v or "") for i, v in enumerate(row)}

    def _iter_rows_xlsx(self, content):
        """Yield dict rows for .xlsx using header names when available."""
        if openpyxl is None:
            raise UserError(_("openpyxl is required to read .xlsx files. Install it or use CSV."))
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        headers = None
        for row in ws.iter_rows(values_only=True):
            row_vals = [(c if c is not None else "") for c in row]
            if headers is None and self.has_header:
                headers = [str(h or "").strip() for h in row_vals]
                continue
            if headers is None:
                # No header case: use numeric string keys
                yield {str(i): v for i, v in enumerate(row_vals)}
            else:
                yield {headers[i] if i < len(headers) else str(i): v for i, v in enumerate(row_vals)}

    # --- helpers to normalize row mapping ---
    def _normalize_row(self, rowdict):
        """Return a dict with keys: name, work_email, job_title, work_phone.

        Accepts multiple header spellings and is case-insensitive.
        Extra columns are ignored.
        """
        def get_any(keys, default=""):
            for k in keys:
                for existing in rowdict.keys():
                    if (existing or "").strip().lower() == k:
                        val = rowdict.get(existing) or ""
                        # Cast everything to string-ish and strip
                        return str(val).strip()
            return default

        name = get_any(["name", "employee name", "employee", "nama"])
        email = get_any(["work_email", "work email", "email", "email address", "e-mail"])
        job_title = get_any(["job_title", "job title", "job position", "position", "title"])
        work_phone = get_any(["work_phone", "work phone", "phone", "mobile", "no hp", "no telp", "telephone", "tel"])

        return {
            "name": name,
            "work_email": email,
            "job_title": job_title,
            "work_phone": work_phone,
        }

    def _detect_ext(self):
        name = (self.file_name or "").lower()
        if name.endswith(".csv"):
            return "csv"
        if name.endswith(".xlsx"):
            return "xlsx"
        raise UserError(_("Unsupported file extension. Use .csv or .xlsx"))

    def action_queue_import(self):
        self.ensure_one()
        content = self._decode_file()
        ext = self._detect_ext()
        row_iter = self._iter_rows_csv(content) if ext == "csv" else self._iter_rows_xlsx(content)

        batches = []
        buf = []
        for rowdict in row_iter:
            vals = self._normalize_row(rowdict or {})
            if vals["name"]:
                buf.append(vals)
                if len(buf) >= (self.chunk_size or 500):
                    batches.append(buf)
                    buf = []
        if buf:
            batches.append(buf)

        if not batches:
            raise UserError(_("No valid rows found."))

        Employee = self.env["hr.employee"].with_context(
            import_notify=self.notify_done,
            import_total_chunks=len(batches),
        )

        delayables = [Employee.delayable()._job_create_employees_from_rows(batch) for batch in batches]
        job_graph = group(*delayables)
        if self.notify_done:
            job_graph = chain(job_graph, Employee.delayable()._job_notify_import_done())
        job_graph.delay()

        action_queue_jobs = self.env.ref("queue_job.action_queue_job").read()[0]

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Employee import queued"),
                "message": _("Queued %s job(s). Track them in Technical → Queue Jobs.") % len(batches),
                "sticky": False,
                "next": action_queue_jobs,
            },
        }